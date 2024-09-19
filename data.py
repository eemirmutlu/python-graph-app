import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

EXCEL_FILE = 'gsyh_data.xlsx'

def save_data_to_excel(country, gsyh):
    data = {
        "Ülke": [country],
        "2020": [gsyh[0]],
        "2021": [gsyh[1]],
        "2022": [gsyh[2]],
        "2023": [gsyh[3]],
    }
    
    df = pd.DataFrame(data)
    
    try:
        book = load_workbook(EXCEL_FILE)
        if 'GSYH_Data' in book.sheetnames:
            df_existing = pd.read_excel(EXCEL_FILE, sheet_name='GSYH_Data')
            
            df_existing = df_existing[df_existing['Ülke'] != country]
            
            df_combined = pd.concat([df_existing, df], ignore_index=True)
            
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer:
                df_combined.to_excel(writer, index=False, sheet_name='GSYH_Data')
        else:
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, index=False, sheet_name='GSYH_Data')
    
    except FileNotFoundError:
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name='GSYH_Data')

def load_country_data(country):
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name='GSYH_Data')
        country_data = df[df['Ülke'] == country]
        if not country_data.empty:
            return country_data.iloc[0, 1:].tolist()
        return None
    except FileNotFoundError:
        return None

def plot_graphs():
    df = pd.read_excel(EXCEL_FILE, sheet_name='GSYH_Data')
    
    countries = df['Ülke'].tolist()
    years = ['2020', '2021', '2022', '2023']
    
    plt.figure(figsize=(18, 8))
    
    bar_width = 0.5
    
    for i, year in enumerate(years):
        plt.subplot(2, 2, i+1)
        bars = plt.bar(countries, df[year], color='blue', width=bar_width)
        plt.title(f'{year} GSYH', fontsize=14)
        plt.xlabel('Ülke', fontsize=10)
        plt.ylabel('GSYH', fontsize=10)
        
        plt.xticks(rotation=45, ha='right', fontsize=8)
        
        for bar in bars:
            yval = bar.get_height()
            plt.text(
                bar.get_x() + bar.get_width() / 2, 
                yval, 
                f'${yval:,.0f}',
                ha='center', 
                va='bottom',
                fontsize=10
            )
    
    plt.tight_layout()
    plt.savefig('gsyh_graphs.png')
    plt.show()
